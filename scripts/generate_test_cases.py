"""Generate a formatted test cases Excel workbook from the test-cases.md source.

Usage:
    python scripts/generate_test_cases.py

Output: docs/test/知识问答子系统测试用例-v0.1.xlsx
"""

import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

DOCS_TEST = Path(__file__).resolve().parent.parent / "docs" / "test"
MD_SOURCE = DOCS_TEST / "test-cases.md"
OUTPUT = DOCS_TEST / "知识问答子系统测试用例-v0.1.xlsx"

# ── colour palette ──────────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_FONT = Font(name="微软雅黑", bold=True, color="FFFFFF", size=10)
CASE_FILL = PatternFill("solid", fgColor="D6E4F0")
BODY_FONT = Font(name="微软雅黑", size=10)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
WRAP = Alignment(wrap_text=True, vertical="top")


def parse_cases(md: str) -> list[dict]:
    """Extract test-case blocks from the markdown document."""
    cases = []
    # Split on "| **用例编号** |" — the split consumes the key,
    # so the captured value (e.g. "TC-HEALTH-001") appears at the
    # start of each block. We re-add the "用例编号" key manually.
    blocks = re.split(r"^\| \*\*用例编号\*\* \|", md.strip(), flags=re.MULTILINE)
    for block in blocks[1:]:
        case: dict[str, str] = {}
        # The first line of the block is the case id value
        lines = block.strip().split("\n")
        if lines:
            first_val = lines[0].strip().rstrip("|").strip()
            case["用例编号"] = first_val
        # Extract remaining key-value pairs
        for m in re.finditer(r"\*\*(.+?)\*\*\s*\|\s*(.+)", block):
            key = m.group(1).strip()
            val = m.group(2).strip().rstrip("|").strip()
            case[key] = val
        cases.append(case)
    return cases


def generate() -> None:
    md = MD_SOURCE.read_text(encoding="utf-8")
    cases = parse_cases(md)

    wb = Workbook()
    ws = wb.active
    ws.title = "测试用例"

    headers = [
        "用例编号", "模块", "接口", "测试名称",
        "前置条件", "测试步骤", "预期结果", "备注",
    ]
    col_widths = [16, 12, 36, 24, 28, 40, 40, 24]

    for i, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 2
    for case in cases:
        values = [
            case.get("用例编号", ""),
            case.get("模块", ""),
            case.get("接口", ""),
            case.get("测试名称", ""),
            case.get("前置条件", ""),
            case.get("测试步骤", ""),
            case.get("预期结果", ""),
            case.get("备注", ""),
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = BODY_FONT
            cell.alignment = WRAP
            cell.border = THIN_BORDER
        row += 1

    DOCS_TEST.mkdir(parents=True, exist_ok=True)
    wb.save(str(OUTPUT))
    print(f"Excel generated: {OUTPUT} ({len(cases)} test cases)")


if __name__ == "__main__":
    generate()
